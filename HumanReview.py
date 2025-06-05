import openpyxl
import tkinter as tk
from tkinter import messagebox
import os
import webbrowser

model = "gpt-4o"

input_file = f"./Output/Search/testcase_search_{model}.xlsx"
output_file = f"./Output/Human/results_{model}.xlsx"
start_index = 0

wb = openpyxl.load_workbook(input_file)
ws = wb.active
row_data_list = list(ws.iter_rows(min_row=2 + start_index, values_only=True))

new_wb = openpyxl.Workbook()
new_ws = new_wb.active
headers = [cell.value for cell in ws[1]]
headers.append("Decision")
new_ws.append(headers)

root = tk.Tk()
root.title("Yes/No Decision Tool")
root.geometry("1000x500")

current_index = 0

progress_label = tk.Label(root, text="", font=("Arial", 14))
progress_label.pack()

display_frame = tk.Frame(root)
display_frame.pack(pady=30)

# Attribute 行
attr_title_label = tk.Label(display_frame, text="Attribute:", font=("Arial", 16, "bold"))
attr_title_label.grid(row=0, column=0, sticky="w", padx=10)

attr_value_entry = tk.Entry(display_frame, font=("Arial", 16), width=80)
attr_value_entry.grid(row=0, column=1, sticky="w", padx=10)

# Content 行
content_title_label = tk.Label(display_frame, text="Content:", font=("Arial", 16, "bold"))
content_title_label.grid(row=1, column=0, sticky="w", padx=10)

content_value_label = tk.Label(display_frame, text="", font=("Arial", 16), wraplength=800, justify="left")
content_value_label.grid(row=1, column=1, sticky="w", padx=10)

# Search num 行
search_label = tk.Label(display_frame, text="Search num:", font=("Arial", 16, "bold"))
search_label.grid(row=2, column=0, sticky="w", padx=10)

search_value_label = tk.Label(display_frame, text="", font=("Arial", 16))
search_value_label.grid(row=2, column=1, sticky="w", padx=10)

def update_question():
    if current_index < len(row_data_list):
        row = row_data_list[current_index]
        attr = str(row[0]) if len(row) > 0 else ""
        content = str(row[1]) if len(row) > 1 else ""
        search_num = str(row[2]) if len(row) > 2 else ""

        attr_value_entry.delete(0, tk.END)
        attr_value_entry.insert(0, attr)
        content_value_label.config(text=content)
        search_value_label.config(text=search_num)

        progress_label.config(text=f"当前进度：{current_index + 1} / {len(row_data_list)}")
    else:
        messagebox.showinfo("完成", f"所有行均已标注完成，结果保存至 {output_file}")
        new_wb.save(output_file)
        root.quit()

def mark(decision):
    global current_index
    if current_index >= len(row_data_list):
        return

    row_data = list(row_data_list[current_index])
    row_data[0] = attr_value_entry.get().strip()
    row_data.append(decision)
    new_ws.append(row_data)

    current_index += 1
    update_question()

def go_back():
    global current_index
    if current_index == 0:
        messagebox.showinfo("提示", "已经是第一条了")
        return

    current_index -= 1
    new_ws.delete_rows(new_ws.max_row, 1)
    update_question()

def open_browser():
    if current_index < len(row_data_list):
        row = row_data_list[current_index]
        content = str(row[1]) if len(row) > 1 else ""
        query = content.strip().replace(" ", "+")
        url = f"https://github.com/search?q={query}&type=code"
        webbrowser.open(url)

btn_frame = tk.Frame(root)
btn_frame.pack(pady=20)

yes_button = tk.Button(btn_frame, text="✅ Yes", command=lambda: mark("Yes"),
                       font=("Arial", 14), bg="lightgreen", width=10)
yes_button.grid(row=0, column=0, padx=10)

no_button = tk.Button(btn_frame, text="❌ No", command=lambda: mark("No"),
                      font=("Arial", 14), bg="lightcoral", width=10)
no_button.grid(row=0, column=1, padx=10)

back_button = tk.Button(btn_frame, text="⬅️ 上一题", command=go_back,
                        font=("Arial", 14), width=10)
back_button.grid(row=0, column=2, padx=10)

browser_button = tk.Button(btn_frame, text="🔎 GitHub搜索", command=open_browser,
                           font=("Arial", 14), width=12)
browser_button.grid(row=0, column=3, padx=10)

update_question()
root.mainloop()